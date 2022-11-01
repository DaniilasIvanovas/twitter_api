import openpyxl


def create_sheet():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Tweets'
    parameters = ['Tweet Id', 'Language', 'Text lenght', 'Hastags', 'Symbols', 'Mentions', 'Url']
    col = 1
    for parameter in parameters:
        cell = sheet.cell(row=1, column=col)
        cell.value = parameter
        col += 1
    wb.save('Duome.xlsx')

create_sheet()
